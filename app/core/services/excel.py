from typing import List

from app.core.models.models import User, Distribution
from app.core.repositories import users as users_repository
from app.core.repositories import distributions as distributions_repository
from app import scheduler

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Border, Side, Alignment


def generate_users_excel_file(users: List[User]) -> str:
    wb = Workbook()
    ws: Worksheet = wb.active
    ws.title = "Пользователи"
    ws.append(["", "Id", "Логин", "Имя", "Телефон", "Добавлен в базу", "Канал", "Рассылка", "Дата рассылки", "Шаблон", "Вышел с канала"])

    for index, user in enumerate(users):
        ws.append([
            index + 1,
            user.telegram_user_id,
            user.username if user.username else 'Не определен',
            user.name,
            user.phone,
            user.created_at,
            user.channel.name if user.channel else "Не определено",
            "",
            "",
            user.distribution.name if user.distribution else "Не определено",
            ""
        ])
    header_sells = ws["A1:K1"]
    thin = Side(border_style='thin', color="000000")
    for cell in header_sells:
        cell[0].border = Border(
            top=thin,
            left=thin,
            right=thin,
            bottom=thin
        )
        cell[0].alignment = Alignment(horizontal="center", vertical="center")
    indexes_cells = ws[f"A1:A{len(users) + 1}"]
    for cell in indexes_cells:
        cell[0].border = Border(
            top=thin,
            left=thin,
            right=thin,
            bottom=thin
        )
        cell[0].alignment = Alignment(horizontal="center", vertical="center")

    path = "storage/users.xlsx"
    wb.save(path)

    return path


def import_excel_file(path: str):
    wb = load_workbook(filename=path)
    ws: Worksheet = wb.active
    for row in ws.iter_rows():
        if not row[0].value:
            continue
        telegram_id = row[1].value
        distribution_template_name = row[9].value

        distribution = distributions_repository.find_distribution_by_name(distribution_template_name)
        users_repository.update_user_distribution(telegram_id, distribution)
        if distribution:
            scheduler.add_distribution(distribution)
